# app/main.py
from __future__ import annotations

from fastapi import FastAPI, Body, Depends, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse

from typing import Dict, Any, List, Optional
from io import BytesIO
from urllib.parse import quote
import datetime as dt
import sqlite3

from sqlmodel import Session, select
from app.database import init_db, get_session, engine, DB_PATH
from app import models

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ---------------------- App & CORS ----------------------
app = FastAPI(title="Alsami Workshop Backend", version="4.0-final")

ALLOWED_ORIGINS = [
    "https://workshop-frontend-cmqd.onrender.com",
    "http://localhost",
    "http://127.0.0.1:5173",
    "http://127.0.0.1:9000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False,   # "*" مع credentials ممنوع؛ نغلقها لتفعيل CORS صحيح
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------- Utilities ----------------------
def _now_iso() -> str:
    return dt.datetime.utcnow().isoformat()


def _pragma_table_info(conn: sqlite3.Connection, table: str) -> List[str]:
    cur = conn.execute(f'PRAGMA table_info("{table}")')
    return [row[1] for row in cur.fetchall()]


def _ensure_table(conn: sqlite3.Connection, table: str):
    """Ensure table exists; create minimal if totally missing."""
    try:
        conn.execute(f'SELECT 1 FROM "{table}" LIMIT 1')
    except sqlite3.OperationalError:
        minimal = {
            "enginesupply":       'CREATE TABLE IF NOT EXISTS enginesupply (id INTEGER PRIMARY KEY, serial TEXT)',
            "engineissue":        'CREATE TABLE IF NOT EXISTS engineissue (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginerehab":        'CREATE TABLE IF NOT EXISTS enginerehab (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginecheck":        'CREATE TABLE IF NOT EXISTS enginecheck (id INTEGER PRIMARY KEY, serial TEXT)',
            "engineupload":       'CREATE TABLE IF NOT EXISTS engineupload (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginelathe":        'CREATE TABLE IF NOT EXISTS enginelathe (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginepump":         'CREATE TABLE IF NOT EXISTS enginepump (id INTEGER PRIMARY KEY, serial TEXT)',
            "engineelectrical":   'CREATE TABLE IF NOT EXISTS engineelectrical (id INTEGER PRIMARY KEY, serial TEXT)',
            "generatorsupply":    'CREATE TABLE IF NOT EXISTS generatorsupply (id INTEGER PRIMARY KEY, code TEXT)',
            "generatorissue":     'CREATE TABLE IF NOT EXISTS generatorissue (id INTEGER PRIMARY KEY, code TEXT)',
            "generatorinspect":   'CREATE TABLE IF NOT EXISTS generatorinspect (id INTEGER PRIMARY KEY, code TEXT)',
        }
        if table in minimal:
            conn.execute(minimal[table])
            conn.commit()


def _ensure_column(conn: sqlite3.Connection, table: str, column: str, coltype: str):
    _ensure_table(conn, table)
    cols = _pragma_table_info(conn, table)
    if column not in cols:
        conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "{column}" {coltype}')
        conn.commit()


def _self_heal_schema():
    """
    Non-destructively add missing columns for all tables.
    Also bridges enginecheck.desc vs. description historical mismatch.
    """
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        # Engines - supply
        for col in ("serial", "engineType", "model", "prevSite", "supDate", "supplier", "notes"):
            _ensure_column(conn, "enginesupply", col, "TEXT")

        # Engines - issue
        for col in ("serial", "currSite", "receiver", "requester", "issueDate", "notes"):
            _ensure_column(conn, "engineissue", col, "TEXT")

        # Engines - rehab
        for col in ("serial", "rehabber", "rehabType", "rehabDate", "notes"):
            _ensure_column(conn, "enginerehab", col, "TEXT")

        # Engines - check (support both description/desc)
        for col in ("serial", "inspector", "checkDate", "notes"):
            _ensure_column(conn, "enginecheck", col, "TEXT")
        cols_check = _pragma_table_info(conn, "enginecheck")
        if "desc" not in cols_check:
            _ensure_column(conn, "enginecheck", "desc", "TEXT")
        if "description" not in cols_check:
            _ensure_column(conn, "enginecheck", "description", "TEXT")
        # fill whichever is empty from the other
        conn.execute('UPDATE enginecheck SET description = COALESCE(description, desc)')
        conn.execute('UPDATE enginecheck SET desc = COALESCE(desc, description)')
        conn.commit()

        # Engines - upload/lathe/pump/electrical
        for col in ("serial", "rehabUp", "checkUp", "rehabUpDate", "checkUpDate", "notes"):
            _ensure_column(conn, "engineupload", col, "TEXT")
        for col in ("serial", "lathe", "latheDate", "notes"):
            _ensure_column(conn, "enginelathe", col, "TEXT")
        for col in ("serial", "pumpSerial", "pumpRehab", "notes"):
            _ensure_column(conn, "enginepump", col, "TEXT")
        for col in ("serial", "etype", "starter", "alternator", "edate", "notes"):
            _ensure_column(conn, "engineelectrical", col, "TEXT")

        # Generators
        for col in ("code", "gType", "model", "prevSite", "supDate", "supplier", "vendor", "notes"):
            _ensure_column(conn, "generatorsupply", col, "TEXT")
        for col in ("code", "issueDate", "receiver", "requester", "currSite", "notes"):
            _ensure_column(conn, "generatorissue", col, "TEXT")
        for col in ("code", "inspector", "elecRehab", "rehabDate", "rehabUp", "checkUp", "notes"):
            _ensure_column(conn, "generatorinspect", col, "TEXT")

    except Exception as e:
        print("⚠️ schema self-heal error:", e)
    finally:
        conn.close()


def ensure_schema(session: Session):
    """Create tables via SQLModel then heal missing SQLite columns."""
    try:
        init_db()
        _self_heal_schema()
    except Exception as e:
        print("⚠️ ensure_schema error:", e)


def to_dict(obj: Any) -> Dict[str, Any]:
    if hasattr(obj, "model_dump"):
        return obj.model_dump(exclude_none=True)
    if hasattr(obj, "dict"):
        return obj.dict(exclude_none=True)
    return dict(obj)


def _safe_list(objs):
    return [to_dict(o) for o in objs]


def _safe_date(s: Optional[str]) -> Optional[dt.date]:
    if not s:
        return None
    try:
        return dt.date.fromisoformat(s)
    except Exception:
        return None


# ---------------------- Startup & Global Error Handler ----------------------
@app.on_event("startup")
def _on_startup():
    init_db()
    _self_heal_schema()


@app.exception_handler(Exception)
async def _global_exception_handler(request: Request, exc: Exception):
    # Avoids Uvicorn "content longer than Content-Length" on unhandled errors
    return JSONResponse(status_code=500, content={"ok": False, "error": str(exc)})


# ---------------------- Health ----------------------
@app.get("/api/health")
def health() -> Dict[str, Any]:
    return {"ok": True, "time": _now_iso()}


# ---------------------- Seed ----------------------
@app.get("/api/seed")
def seed(session: Session = Depends(get_session)):
    """Seed minimal demo data if DB is empty."""
    try:
        ensure_schema(session)
        if session.exec(select(models.EngineSupply).limit(1)).first():
            return {"ok": True, "msg": "data-exists"}
        try:
            from app.seed_demo_data import seed as do_seed
            do_seed(session)
        except Exception:
            session.add(models.EngineSupply(
                serial="111", engineType="Deutz", model="F4L912",
                prevSite="المخزن", supDate="2025-10-30",
                supplier="Yemen Mobile", notes="توريد جديد"
            ))
            session.add(models.GeneratorSupply(
                code="GEN001", gType="30kVA", model="FG Wilson",
                prevSite="المستودع", supDate="2025-10-30",
                supplier="Yemen Mobile", vendor="PowerMax", notes="جديد"
            ))
            session.commit()
        return {"ok": True, "msg": "seeded"}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ---------------------- Admin: Repair / Rebuild ----------------------
@app.post("/api/admin/repair")
def admin_repair():
    try:
        _self_heal_schema()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@app.post("/api/admin/rebuild")
def admin_rebuild():
    try:
        init_db()
        _self_heal_schema()
        return {"ok": True, "msg": "rebuilt"}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ---------------------- Search & Last3 ----------------------
@app.get("/api/search/{key}")
def search_item(key: str, session: Session = Depends(get_session)) -> Dict[str, Any]:
    """
    Returns full records for engines (by serial) and generators (by code).
    """
    ensure_schema(session)
    result: Dict[str, Any] = {
        "engines": {
            "supply": [], "issue": [], "rehab": [], "check": [],
            "upload": [], "lathe": [], "pump": [], "electrical": []
        },
        "generators": { "supply": [], "issue": [], "inspect": [] },
    }
    try:
        # Engines
        result["engines"]["supply"] = _safe_list(
            session.exec(select(models.EngineSupply).where(models.EngineSupply.serial == key)).all()
        )
        result["engines"]["issue"] = _safe_list(
            session.exec(select(models.EngineIssue).where(models.EngineIssue.serial == key)).all()
        )
        result["engines"]["rehab"] = _safe_list(
            session.exec(select(models.EngineRehab).where(models.EngineRehab.serial == key)).all()
        )
        result["engines"]["check"] = _safe_list(
            session.exec(select(models.EngineCheck).where(models.EngineCheck.serial == key)).all()
        )
        result["engines"]["upload"] = _safe_list(
            session.exec(select(models.EngineUpload).where(models.EngineUpload.serial == key)).all()
        )
        result["engines"]["lathe"] = _safe_list(
            session.exec(select(models.EngineLathe).where(models.EngineLathe.serial == key)).all()
        )
        result["engines"]["pump"] = _safe_list(
            session.exec(select(models.EnginePump).where(models.EnginePump.serial == key)).all()
        )
        result["engines"]["electrical"] = _safe_list(
            session.exec(select(models.EngineElectrical).where(models.EngineElectrical.serial == key)).all()
        )

        # Generators
        result["generators"]["supply"] = _safe_list(
            session.exec(select(models.GeneratorSupply).where(models.GeneratorSupply.code == key)).all()
        )
        result["generators"]["issue"] = _safe_list(
            session.exec(select(models.GeneratorIssue).where(models.GeneratorIssue.code == key)).all()
        )
        result["generators"]["inspect"] = _safe_list(
            session.exec(select(models.GeneratorInspect).where(models.GeneratorInspect.code == key)).all()
        )
        return result
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"search-error: {e}"}, status_code=500)


@app.get("/api/last3/engines")
def last3_engines(session: Session = Depends(get_session)) -> Dict[str, Any]:
    ensure_schema(session)
    try:
        rows = session.exec(
            select(models.EngineSupply).order_by(models.EngineSupply.id.desc()).limit(3)
        ).all()
        return {"items": [{"serial": r.serial, "prevSite": r.prevSite or ""} for r in rows]}
    except Exception as e:
        return {"items": [], "warn": str(e)}


@app.get("/api/last3/generators")
def last3_generators(session: Session = Depends(get_session)) -> Dict[str, Any]:
    ensure_schema(session)
    try:
        rows = session.exec(
            select(models.GeneratorSupply).order_by(models.GeneratorSupply.id.desc()).limit(3)
        ).all()
        return {"items": [{"code": r.code, "prevSite": r.prevSite or ""} for r in rows]}
    except Exception as e:
        return {"items": [], "warn": str(e)}


# ---------------------- Sync (batch) & Save+Search ----------------------
@app.post("/api/sync/batch")
def sync_batch(payload: Dict[str, Any] = Body(...), session: Session = Depends(get_session)):
    """
    payload = { items: [ {store: 'eng_supply' | 'gen_issue' | ... , payload: {..fields..}} ] }
    Returns {"ok":True,"saved":..,"skipped":..}
    """
    ensure_schema(session)
    items: List[Dict[str, Any]] = payload.get("items", [])
    saved, skipped = 0, 0

    STORE_TO_MODEL = {
        # Engines
        "eng_supply": models.EngineSupply,
        "eng_issue": models.EngineIssue,
        "eng_rehab": models.EngineRehab,
        "eng_check": models.EngineCheck,
        "eng_upload": models.EngineUpload,
        "eng_lathe": models.EngineLathe,
        "eng_pump": models.EnginePump,
        "eng_electrical": models.EngineElectrical,
        # Generators
        "gen_supply": models.GeneratorSupply,
        "gen_issue": models.GeneratorIssue,
        "gen_inspect": models.GeneratorInspect,
    }

    for it in items:
        try:
            store = it.get("store")
            data = it.get("payload", {}) or {}
            Model = STORE_TO_MODEL.get(store)
            if not Model:
                skipped += 1
                continue
            obj = Model(**data)
            session.add(obj)
            saved += 1
        except Exception:
            skipped += 1

    session.commit()
    return {"ok": True, "saved": saved, "skipped": skipped}


@app.post("/api/save-and-search")
def save_and_search(payload: Dict[str, Any] = Body(...), session: Session = Depends(get_session)):
    """
    Save single item then return merged search result.
    payload = { store: 'eng_supply' | ... , payload: {...}, key: '111' | 'GEN001' }
    """
    ensure_schema(session)
    try:
        STORE_TO_MODEL = {
            "eng_supply": models.EngineSupply,
            "eng_issue": models.EngineIssue,
            "eng_rehab": models.EngineRehab,
            "eng_check": models.EngineCheck,
            "eng_upload": models.EngineUpload,
            "eng_lathe": models.EngineLathe,
            "eng_pump": models.EnginePump,
            "eng_electrical": models.EngineElectrical,
            "gen_supply": models.GeneratorSupply,
            "gen_issue": models.GeneratorIssue,
            "gen_inspect": models.GeneratorInspect,
        }
        store = payload.get("store")
        data = payload.get("payload", {}) or {}
        key = payload.get("key") or data.get("serial") or data.get("code")
        Model = STORE_TO_MODEL.get(store)
        if not Model:
            raise HTTPException(status_code=400, detail="invalid store")
        session.add(Model(**data))
        session.commit()
        return search_item(str(key), session)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"save-and-search error: {e}")


# ---------------------- Export (aggregated rows) ----------------------
@app.post("/api/export/xlsx")
def export_xlsx(payload: Dict[str, Any] = Body(...), session: Session = Depends(get_session)):
    """
    body:
    {
      "scope": "engines"|"generators"|"both",
      "date_from": "YYYY-MM-DD",   # اختياري
      "date_to":   "YYYY-MM-DD",   # اختياري
      "filename": "تقرير.xlsx"     # اختياري
    }

    مخرجات:
    - شيت للمحركات: صف واحد لكل رقم تسلسلي مع أعمدة منفصلة لكل نوع حركة.
    - شيت للمولدات: صف واحد لكل كود مع أعمدة منفصلة لكل نوع حركة.
    - لو فيه أكثر من حركة لنفس الرقم/الكود في نفس النوع، يتم دمج القيم في نفس العمود مفصولة بـ " | ".
    """
    ensure_schema(session)

    scope = (payload.get("scope") or "both").lower()
    date_from = _safe_date(payload.get("date_from"))
    date_to = _safe_date(payload.get("date_to"))
    original_filename: str = payload.get("filename") or "report.xlsx"
    safe_filename = "report.xlsx"
    encoded_name = quote(original_filename)

    def in_range(date_str: Optional[str]) -> bool:
        if not (date_from or date_to):
            return True
        d = _safe_date(date_str)
        if not d:
            return False
        if date_from and d < date_from:
            return False
        if date_to and d > date_to:
            return False
        return True

    # إعداد ملف Excel والتنسيقات
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="2563eb")
    alt_fill = PatternFill("solid", fgColor="f1f5f9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    normal_font = Font(name="Arial")
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def set_or_join(row: Dict[str, str], key: str, value: str):
        """يجمع القيم المتعددة لنفس الحقل في نفس العمود."""
        if not value:
            return
        old = row.get(key)
        if not old:
            row[key] = value
        elif value not in old.split(" | "):
            row[key] = f"{old} | {value}"

    # ===================== Sheet المحركات =====================
    ws_eng = None
    if scope in ("engines", "both"):
        ws_eng = wb.active
        ws_eng.title = "المحركات"
        ws_eng.sheet_view.rightToLeft = True

        eng_headers = [
            "الرقم التسلسلي",

            # توريد
            "توريد - النوع",
            "توريد - الموديل",
            "توريد - الموقع السابق",
            "توريد - تاريخ التوريد",
            "توريد - المورد",
            "توريد - ملاحظات",

            # صرف
            "صرف - الموقع الحالي",
            "صرف - المستلم",
            "صرف - طالب الصرف",
            "صرف - تاريخ الصرف",
            "صرف - ملاحظات",

            # تأهيل
            "تأهيل - الجهة",
            "تأهيل - نوع التأهيل",
            "تأهيل - تاريخ التأهيل",
            "تأهيل - ملاحظات",

            # فحص
            "فحص - الفاحص",
            "فحص - الوصف",
            "فحص - تاريخ الفحص",
            "فحص - ملاحظات",

            # رفع
            "رفع - رفع تأهيل",
            "رفع - رفع فحص",
            "رفع - تاريخ رفع التأهيل",
            "رفع - تاريخ رفع الفحص",
            "رفع - ملاحظات",

            # مخرطة
            "مخرطة - التفاصيل",
            "مخرطة - تاريخ المخرطة",
            "مخرطة - ملاحظات",

            # بمبات ونوزلات
            "بمبات/نوزلات - رقم البمب",
            "بمبات/نوزلات - تفاصيل التأهيل",
            "بمبات/نوزلات - ملاحظات",

            # كهرباء
            "كهرباء - النوع",
            "كهرباء - السلف",
            "كهرباء - الدينمو",
            "كهرباء - تاريخ",
            "كهرباء - ملاحظات",
        ]

        ws_eng.append(eng_headers)
        for i in range(1, len(eng_headers) + 1):
            c = ws_eng.cell(row=1, column=i)
            c.fill = header_fill
            c.font = header_font
            c.alignment = align_right
            c.border = thin_border
            ws_eng.column_dimensions[c.column_letter].width = 25

        supplies = session.exec(select(models.EngineSupply)).all()
        issues   = session.exec(select(models.EngineIssue)).all()
        rehabs   = session.exec(select(models.EngineRehab)).all()
        checks   = session.exec(select(models.EngineCheck)).all()
        uploads  = session.exec(select(models.EngineUpload)).all()
        lathes   = session.exec(select(models.EngineLathe)).all()
        pumps    = session.exec(select(models.EnginePump)).all()
        elecs    = session.exec(select(models.EngineElectrical)).all()

        by_serial: Dict[str, Dict[str, str]] = {}

        # توريد
        for r in supplies:
            if not in_range(r.supDate):
                continue
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            set_or_join(row, "sup_type",      r.engineType or "")
            set_or_join(row, "sup_model",     r.model or "")
            set_or_join(row, "sup_prevSite",  r.prevSite or "")
            set_or_join(row, "sup_date",      r.supDate or "")
            set_or_join(row, "sup_supplier",  r.supplier or "")
            set_or_join(row, "sup_notes",     r.notes or "")

        # صرف
        for r in issues:
            if not in_range(r.issueDate):
                continue
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            set_or_join(row, "iss_currSite",  r.currSite or "")
            set_or_join(row, "iss_receiver",  r.receiver or "")
            set_or_join(row, "iss_requester", r.requester or "")
            set_or_join(row, "iss_date",      r.issueDate or "")
            set_or_join(row, "iss_notes",     r.notes or "")

        # تأهيل
        for r in rehabs:
            if not in_range(r.rehabDate):
                continue
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            set_or_join(row, "reh_rehabber",  r.rehabber or "")
            set_or_join(row, "reh_type",      r.rehabType or "")
            set_or_join(row, "reh_date",      r.rehabDate or "")
            set_or_join(row, "reh_notes",     r.notes or "")

        # فحص
        for r in checks:
            if not in_range(r.checkDate):
                continue
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            desc = getattr(r, "description", None) or getattr(r, "desc", "") or ""
            set_or_join(row, "chk_inspector", r.inspector or "")
            set_or_join(row, "chk_desc",      desc)
            set_or_join(row, "chk_date",      r.checkDate or "")
            set_or_join(row, "chk_notes",     r.notes or "")

        # رفع
        for r in uploads:
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            set_or_join(row, "upl_rehabUp",     r.rehabUp or "")
            set_or_join(row, "upl_checkUp",     r.checkUp or "")
            set_or_join(row, "upl_rehabUpDate", r.rehabUpDate or "")
            set_or_join(row, "upl_checkUpDate", r.checkUpDate or "")
            set_or_join(row, "upl_notes",       r.notes or "")

        # مخرطة
        for r in lathes:
            if not in_range(r.latheDate):
                continue
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            set_or_join(row, "lathe_name",  r.lathe or "")
            set_or_join(row, "lathe_date",  r.latheDate or "")
            set_or_join(row, "lathe_notes", r.notes or "")

        # بمبات ونوزلات
        for r in pumps:
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            set_or_join(row, "pump_serial", r.pumpSerial or "")
            set_or_join(row, "pump_rehab",  r.pumpRehab or "")
            set_or_join(row, "pump_notes",  r.notes or "")

        # كهرباء
        for r in elecs:
            if not in_range(r.edate):
                continue
            row = by_serial.setdefault(r.serial, {"serial": r.serial})
            set_or_join(row, "elec_type",   r.etype or "")
            set_or_join(row, "elec_start",  r.starter or "")
            set_or_join(row, "elec_alt",    r.alternator or "")
            set_or_join(row, "elec_date",   r.edate or "")
            set_or_join(row, "elec_notes",  r.notes or "")

        # كتابة الصفوف
        row_idx = 2
        for serial, segs in sorted(by_serial.items()):
            row_vals = [
                serial,

                segs.get("sup_type", ""),
                segs.get("sup_model", ""),
                segs.get("sup_prevSite", ""),
                segs.get("sup_date", ""),
                segs.get("sup_supplier", ""),
                segs.get("sup_notes", ""),

                segs.get("iss_currSite", ""),
                segs.get("iss_receiver", ""),
                segs.get("iss_requester", ""),
                segs.get("iss_date", ""),
                segs.get("iss_notes", ""),

                segs.get("reh_rehabber", ""),
                segs.get("reh_type", ""),
                segs.get("reh_date", ""),
                segs.get("reh_notes", ""),

                segs.get("chk_inspector", ""),
                segs.get("chk_desc", ""),
                segs.get("chk_date", ""),
                segs.get("chk_notes", ""),

                segs.get("upl_rehabUp", ""),
                segs.get("upl_checkUp", ""),
                segs.get("upl_rehabUpDate", ""),
                segs.get("upl_checkUpDate", ""),
                segs.get("upl_notes", ""),

                segs.get("lathe_name", ""),
                segs.get("lathe_date", ""),
                segs.get("lathe_notes", ""),

                segs.get("pump_serial", ""),
                segs.get("pump_rehab", ""),
                segs.get("pump_notes", ""),

                segs.get("elec_type", ""),
                segs.get("elec_start", ""),
                segs.get("elec_alt", ""),
                segs.get("elec_date", ""),
                segs.get("elec_notes", ""),
            ]
            ws_eng.append(row_vals)
            for col in range(1, len(row_vals) + 1):
                c = ws_eng.cell(row=row_idx, column=col)
                c.font = normal_font
                c.alignment = align_right
                c.border = thin_border
                c.fill = alt_fill if row_idx % 2 else white_fill
            row_idx += 1

    # ===================== Sheet المولدات =====================
    ws_gen = None
    if scope in ("generators", "both"):
        ws_gen = wb.create_sheet(title="المولدات") if ws_eng else wb.active
        ws_gen.title = "المولدات"
        ws_gen.sheet_view.rightToLeft = True

        gen_headers = [
            "كود المولد",

            # توريد
            "توريد - السعة",
            "توريد - الموديل",
            "توريد - الموقع السابق",
            "توريد - تاريخ التوريد",
            "توريد - الجهة",
            "توريد - المورد",
            "توريد - ملاحظات",

            # صرف
            "صرف - تاريخ الصرف",
            "صرف - المستلم",
            "صرف - طالب الصرف",
            "صرف - الموقع الحالي",
            "صرف - ملاحظات",

            # فحص/رفع
            "فحص/رفع - المفتش",
            "فحص/رفع - تأهيل كهربائي",
            "فحص/رفع - تاريخ التأهيل",
            "فحص/رفع - رفع تأهيل",
            "فحص/رفع - رفع فحص",
            "فحص/رفع - ملاحظات",
        ]
        ws_gen.append(gen_headers)
        for i in range(1, len(gen_headers) + 1):
            c = ws_gen.cell(row=1, column=i)
            c.fill = header_fill
            c.font = header_font
            c.alignment = align_right
            c.border = thin_border
            ws_gen.column_dimensions[c.column_letter].width = 25

        sup = session.exec(select(models.GeneratorSupply)).all()
        iss = session.exec(select(models.GeneratorIssue)).all()
        ins = session.exec(select(models.GeneratorInspect)).all()

        by_code: Dict[str, Dict[str, str]] = {}

        # توريد
        for r in sup:
            if not in_range(r.supDate):
                continue
            row = by_code.setdefault(r.code, {"code": r.code})
            set_or_join(row, "sup_gType",    r.gType or "")
            set_or_join(row, "sup_model",    r.model or "")
            set_or_join(row, "sup_prevSite", r.prevSite or "")
            set_or_join(row, "sup_date",     r.supDate or "")
            set_or_join(row, "sup_supplier", r.supplier or "")
            set_or_join(row, "sup_vendor",   r.vendor or "")
            set_or_join(row, "sup_notes",    r.notes or "")

        # صرف
        for r in iss:
            if not in_range(r.issueDate):
                continue
            row = by_code.setdefault(r.code, {"code": r.code})
            set_or_join(row, "iss_date",      r.issueDate or "")
            set_or_join(row, "iss_receiver",  r.receiver or "")
            set_or_join(row, "iss_requester", r.requester or "")
            set_or_join(row, "iss_currSite",  r.currSite or "")
            set_or_join(row, "iss_notes",     r.notes or "")

        # فحص/رفع
        for r in ins:
            if not in_range(r.rehabDate):
                # لو مافي rehabDate نعتبره داخل النطاق
                if date_from or date_to:
                    continue
            row = by_code.setdefault(r.code, {"code": r.code})
            set_or_join(row, "ins_inspector", r.inspector or "")
            set_or_join(row, "ins_elecRehab", r.elecRehab or "")
            set_or_join(row, "ins_rehabDate", r.rehabDate or "")
            set_or_join(row, "ins_rehabUp",   r.rehabUp or "")
            set_or_join(row, "ins_checkUp",   r.checkUp or "")
            set_or_join(row, "ins_notes",     r.notes or "")

        # كتابة الصفوف
        row_idx = 2
        for code, segs in sorted(by_code.items()):
            row_vals = [
                code,

                segs.get("sup_gType", ""),
                segs.get("sup_model", ""),
                segs.get("sup_prevSite", ""),
                segs.get("sup_date", ""),
                segs.get("sup_supplier", ""),
                segs.get("sup_vendor", ""),
                segs.get("sup_notes", ""),

                segs.get("iss_date", ""),
                segs.get("iss_receiver", ""),
                segs.get("iss_requester", ""),
                segs.get("iss_currSite", ""),
                segs.get("iss_notes", ""),

                segs.get("ins_inspector", ""),
                segs.get("ins_elecRehab", ""),
                segs.get("ins_rehabDate", ""),
                segs.get("ins_rehabUp", ""),
                segs.get("ins_checkUp", ""),
                segs.get("ins_notes", ""),
            ]
            ws_gen.append(row_vals)
            for col in range(1, len(row_vals) + 1):
                c = ws_gen.cell(row=row_idx, column=col)
                c.font = normal_font
                c.alignment = align_right
                c.border = thin_border
                c.fill = alt_fill if row_idx % 2 else white_fill
            row_idx += 1

    # لو لا محركات ولا مولدات
    if not ws_eng and not ws_gen:
        ws = wb.active
        ws.title = "فارغ"
        ws.append(["لا توجد بيانات للتصدير"])

    # توقيع
    ws = ws_gen or ws_eng or wb.active
    sig = ws.cell(row=ws.max_row + 2, column=1)
    sig.value = f"تاريخ التوليد: {dt.datetime.utcnow().isoformat()}"
    sig.alignment = align_right
    sig.font = Font(italic=True, color="666666", name="Arial")

    # تجهيز الاستجابة
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers_resp = {
        "Content-Disposition": f"attachment; filename={safe_filename}; filename*=UTF-8''{encoded_name}",
        "Access-Control-Expose-Headers": "Content-Disposition",
    }

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


# ---------------------- Root & Favicon ----------------------
@app.get("/")
def root():
    return {"service": "Workshop API", "status": "running", "static": "/static/index.html"}


@app.get("/favicon.ico")
def favicon():
    return JSONResponse(content={}, status_code=204)


# ---------------------- Local run ----------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=9000, reload=True)
